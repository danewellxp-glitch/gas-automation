"""API do módulo Financeiro."""
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import and_, func, select
from sqlalchemy.exc import ProgrammingError
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import get_current_user
from app.database import get_db
from app.models.auth_models import User
from app.models.financeiro.account import Account
from app.models.financeiro.transaction import Transaction, TransactionType
from app.models.financeiro.receivable import Receivable, ReceivableStatus
from app.models.financeiro.payable import Payable, PayableStatus
from app.models.financeiro.cost_center import CostCenter
from app.schemas.financeiro import (
    AccountCreate, AccountResponse, AccountUpdate,
    CostCenterCreate, CostCenterResponse,
    FinanceiroDashboardResponse,
    PayableCreate, PayableResponse,
    ReceivableCreate, ReceivableResponse,
    TransactionCreate, TransactionResponse, TransactionUpdate,
    PaginatedTransactionsResponse,
    DREResponse, CashFlowResponse,
)
from app.services.financeiro.financial_service import (
    create_transaction as svc_create_transaction,
    get_dashboard,
    get_dre,
    get_cash_flow,
)

router = APIRouter(tags=["Financeiro"])

ALLOWED_ROLES = ["financeiro", "admin", "owner", "operator"]


def _require_financeiro(user: User) -> None:
    if user.role not in ALLOWED_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acesso negado")


# ---- Dashboard ----

@router.get("/dashboard", response_model=FinanceiroDashboardResponse)
async def financeiro_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    return await get_dashboard(db)


# ---- Accounts ----

@router.get("/accounts", response_model=List[AccountResponse])
async def list_accounts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(
        select(Account).where(Account.deleted_at == None).order_by(Account.name)
    )
    return result.scalars().all()


@router.post("/accounts", response_model=AccountResponse, status_code=201)
async def create_account(
    body: AccountCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    account = Account(
        name=body.name,
        type=body.type,
        bank_name=body.bank_name,
        agency=body.agency,
        account_number=body.account_number,
        balance=body.initial_balance,
        initial_balance=body.initial_balance,
    )
    db.add(account)
    await db.commit()
    await db.refresh(account)
    return account


@router.put("/accounts/{account_id}", response_model=AccountResponse)
async def update_account(
    account_id: UUID,
    body: AccountUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Account).where(Account.id == account_id, Account.deleted_at == None))
    account = result.scalar_one_or_none()
    if not account:
        raise HTTPException(404, "Conta não encontrada")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(account, field, value)
    await db.commit()
    await db.refresh(account)
    return account


# ---- Transactions ----

@router.get("/transactions", response_model=PaginatedTransactionsResponse)
async def list_transactions(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    type: Optional[str] = None,
    category: Optional[str] = None,
    account_id: Optional[UUID] = None,
    is_paid: Optional[bool] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    filters = [Transaction.deleted_at == None]
    if type:
        filters.append(Transaction.type == type)
    if category:
        filters.append(Transaction.category == category)
    if account_id:
        filters.append(Transaction.account_id == account_id)
    if is_paid is not None:
        filters.append(Transaction.is_paid == is_paid)
    if start_date:
        filters.append(Transaction.reference_date >= start_date)
    if end_date:
        filters.append(Transaction.reference_date <= end_date)

    count_result = await db.execute(select(func.count(Transaction.id)).where(and_(*filters)))
    total = count_result.scalar() or 0

    offset = (page - 1) * per_page
    result = await db.execute(
        select(Transaction).where(and_(*filters))
        .order_by(Transaction.reference_date.desc(), Transaction.created_at.desc())
        .offset(offset).limit(per_page)
    )
    items = result.scalars().all()
    import math
    return PaginatedTransactionsResponse(
        items=items, total=total, page=page, per_page=per_page,
        pages=math.ceil(total / per_page) if per_page else 1
    )


@router.post("/transactions", response_model=TransactionResponse, status_code=201)
async def create_transaction_endpoint(
    body: TransactionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    tx = await svc_create_transaction(
        db=db,
        account_id=body.account_id,
        type=body.type,
        category=body.category,
        description=body.description,
        amount=body.amount,
        direction=body.direction,
        reference_date=body.reference_date,
        created_by=current_user.id,
        due_date=body.due_date,
        is_paid=body.is_paid,
        order_id=body.order_id,
        cost_center_id=body.cost_center_id,
        notes=body.notes,
    )
    await db.commit()
    await db.refresh(tx)
    return tx


@router.put("/transactions/{tx_id}", response_model=TransactionResponse)
async def update_transaction(
    tx_id: UUID,
    body: TransactionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Transaction).where(Transaction.id == tx_id, Transaction.deleted_at == None))
    tx = result.scalar_one_or_none()
    if not tx:
        raise HTTPException(404, "Transação não encontrada")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(tx, field, value)
    await db.commit()
    await db.refresh(tx)
    return tx


@router.delete("/transactions/{tx_id}", status_code=204)
async def delete_transaction(
    tx_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Transaction).where(Transaction.id == tx_id, Transaction.deleted_at == None))
    tx = result.scalar_one_or_none()
    if not tx:
        raise HTTPException(404, "Transação não encontrada")
    tx.deleted_at = datetime.now(timezone.utc)
    await db.commit()


@router.post("/transactions/{tx_id}/pay", response_model=TransactionResponse)
async def pay_transaction(
    tx_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Transaction).where(Transaction.id == tx_id, Transaction.deleted_at == None))
    tx = result.scalar_one_or_none()
    if not tx:
        raise HTTPException(404, "Transação não encontrada")
    tx.is_paid = True
    tx.paid_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(tx)
    return tx


# ---- Receivables ----

@router.get("/receivables", response_model=List[ReceivableResponse])
async def list_receivables(
    status_filter: Optional[str] = Query(None, alias="status"),
    due_from: Optional[date] = None,
    due_to: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    filters = [Receivable.deleted_at == None]
    if status_filter:
        filters.append(Receivable.status == status_filter)
    if due_from:
        filters.append(Receivable.due_date >= due_from)
    if due_to:
        filters.append(Receivable.due_date <= due_to)
    result = await db.execute(
        select(Receivable).where(and_(*filters)).order_by(Receivable.due_date.asc())
    )
    return result.scalars().all()


@router.post("/receivables", response_model=ReceivableResponse, status_code=201)
async def create_receivable(
    body: ReceivableCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    recv = Receivable(**body.model_dump())
    db.add(recv)
    await db.commit()
    await db.refresh(recv)
    return recv


@router.post("/receivables/{recv_id}/receive", response_model=ReceivableResponse)
async def receive_receivable(
    recv_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Receivable).where(Receivable.id == recv_id, Receivable.deleted_at == None))
    recv = result.scalar_one_or_none()
    if not recv:
        raise HTTPException(404, "Recebível não encontrado")
    recv.status = ReceivableStatus.recebido.value
    recv.received_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(recv)
    return recv


# ---- Payables ----

@router.get("/payables", response_model=List[PayableResponse])
async def list_payables(
    status_filter: Optional[str] = Query(None, alias="status"),
    due_from: Optional[date] = None,
    due_to: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    filters = [Payable.deleted_at == None]
    if status_filter:
        filters.append(Payable.status == status_filter)
    if due_from:
        filters.append(Payable.due_date >= due_from)
    if due_to:
        filters.append(Payable.due_date <= due_to)
    result = await db.execute(
        select(Payable).where(and_(*filters)).order_by(Payable.due_date.asc())
    )
    return result.scalars().all()


@router.post("/payables", response_model=PayableResponse, status_code=201)
async def create_payable(
    body: PayableCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    payable = Payable(**body.model_dump())
    db.add(payable)
    await db.commit()
    await db.refresh(payable)
    return payable


@router.post("/payables/{payable_id}/pay", response_model=PayableResponse)
async def pay_payable(
    payable_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(Payable).where(Payable.id == payable_id, Payable.deleted_at == None))
    payable = result.scalar_one_or_none()
    if not payable:
        raise HTTPException(404, "Conta a pagar não encontrada")
    payable.status = PayableStatus.pago.value
    payable.paid_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(payable)
    return payable


# ---- Cash Flow / DRE ----

@router.get("/cash-flow", response_model=CashFlowResponse)
async def cash_flow(
    days: int = Query(30, ge=7, le=90),
    account_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    return await get_cash_flow(db, account_id, days)


@router.get("/dre", response_model=DREResponse)
async def dre(
    start: date = Query(...),
    end: date = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    return await get_dre(db, start, end)


# ---- Cost Centers ----

@router.get("/cost-centers", response_model=List[CostCenterResponse])
async def list_cost_centers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(select(CostCenter).where(CostCenter.is_active == True))
    return result.scalars().all()


@router.post("/cost-centers", response_model=CostCenterResponse, status_code=201)
async def create_cost_center(
    body: CostCenterCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    cc = CostCenter(**body.model_dump())
    db.add(cc)
    await db.commit()
    await db.refresh(cc)
    return cc


# ---- Reports ----

@router.get("/reports/monthly")
async def monthly_report(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    import calendar
    start = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end = date(year, month, last_day)
    return await get_dre(db, start, end)


@router.get("/reports/by-category")
async def report_by_category(
    start: date = Query(...),
    end: date = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_financeiro(current_user)
    result = await db.execute(
        select(Transaction.category, Transaction.type, func.sum(Transaction.amount))
        .where(and_(
            Transaction.reference_date >= start,
            Transaction.reference_date <= end,
            Transaction.deleted_at == None,
        ))
        .group_by(Transaction.category, Transaction.type)
        .order_by(Transaction.type, func.sum(Transaction.amount).desc())
    )
    rows = result.all()
    return [{"category": r[0], "type": r[1], "total": float(r[2])} for r in rows]


# ─────────────────────────────────────────────────────────
# Insights — lucratividade e fiado
# ─────────────────────────────────────────────────────────

@router.get("/insights/bairros")
async def get_top_bairros(
    limit: int = Query(10, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Top bairros por lucro total."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import get_top_profitable_bairros
    return await get_top_profitable_bairros(db, limit=limit)


@router.get("/insights/expensive-deliveries")
async def get_expensive_deliveries(
    limit: int = Query(10, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Entregas com maior custo estimado (menor margem)."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import get_most_expensive_deliveries
    return await get_most_expensive_deliveries(db, limit=limit)


@router.get("/customers/debt")
async def get_customers_debt(
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista clientes com dívida (balance negativo), ordenados pelo maior devedor."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import get_high_risk_customers
    return await get_high_risk_customers(db, limit=limit)


@router.get("/customers/{customer_id}/financial")
async def get_customer_financial(
    customer_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Situação financeira de um cliente específico."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import get_customer_financial_summary
    summary = await get_customer_financial_summary(db, customer_id)
    if not summary:
        raise HTTPException(status_code=404, detail="Nenhum histórico financeiro para este cliente")
    return summary


@router.post("/customers/{customer_id}/credit-limit")
async def update_credit_limit(
    customer_id: UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Atualiza o limite de crédito (fiado) de um cliente."""
    _require_financeiro(current_user)
    from app.models.financeiro.customer_financial import CustomerFinancial
    from app.services.financeiro.customer_financial_service import _get_or_create_customer_financial

    new_limit = Decimal(str(body.get("credit_limit", 200)))
    if new_limit < 0:
        raise HTTPException(status_code=400, detail="Limite não pode ser negativo")

    cf = await _get_or_create_customer_financial(db, customer_id)
    cf.credit_limit = new_limit
    await db.commit()
    return {"customer_id": str(customer_id), "credit_limit": float(cf.credit_limit)}


@router.get("/orders/profit")
async def get_orders_profit(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lucro por pedido, do mais recente ao mais antigo."""
    _require_financeiro(current_user)
    from app.models.financeiro.order_profit import OrderProfit
    from app.models.order import Order

    result = await db.execute(
        select(OrderProfit, Order)
        .join(Order, OrderProfit.order_id == Order.id)
        .order_by(OrderProfit.calculated_at.desc())
        .limit(limit)
        .offset(offset)
    )
    rows = result.all()
    return [
        {
            "order_id":        str(op.order_id),
            "order_number":    order.order_number,
            "bairro":          op.bairro,
            "revenue":         float(op.revenue),
            "product_cost":    float(op.product_cost),
            "delivery_cost":   float(op.delivery_cost),
            "estimated_cost":  float(op.estimated_cost),
            "profit":          float(op.profit),
            "margin_percentage": float(op.margin_percentage),
            "calculated_at":   op.calculated_at.isoformat(),
        }
        for op, order in rows
    ]


@router.post("/orders/{order_id}/recalculate-profit")
async def recalculate_profit(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Força recálculo do lucro de um pedido específico."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import calculate_order_profit
    op = await calculate_order_profit(db, order_id)
    if not op:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    await db.commit()
    return {"order_id": str(order_id), "profit": float(op.profit), "margin": float(op.margin_percentage)}


# ============================================================
# CONCILIAÇÃO PIX
# ============================================================

from pydantic import BaseModel as PydanticBaseModel, ConfigDict, Field as PydanticField


class ConciliacaoRunResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    conciliation_date: date
    run_at: datetime
    total_asaas: int
    total_local: int
    total_conciliado: int
    total_divergencia: int
    total_apenas_asaas: int
    total_apenas_local: int
    valor_asaas: Decimal
    valor_local: Decimal
    status: str


class ConciliacaoItemResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    resultado: str
    asaas_payment_id: Optional[str]
    asaas_amount: Optional[Decimal]
    asaas_payer_name: Optional[str]
    transaction_amount: Optional[Decimal]
    order_id: Optional[UUID]
    diferenca: Decimal
    observacao: Optional[str]
    resolvido: bool


class ConciliacaoRunFullResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    run: ConciliacaoRunResponse
    items: List[ConciliacaoItemResponse]


@router.get("/conciliacao-pix")
async def get_conciliacao_pix(
    date_param: date = Query(None, alias="date"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Busca ou cria conciliação PIX do dia."""
    _require_financeiro(current_user)
    from app.services.financeiro.pix_conciliation_service import PixConciliationService
    from app.models.financeiro.pix_conciliation import PixConciliationItem, PixConciliationRun

    target = date_param or date.today()
    svc = PixConciliationService()
    run = await svc.get_or_run_conciliation(db, target)

    items_result = await db.execute(
        select(PixConciliationItem).where(PixConciliationItem.run_id == run.id)
        .order_by(PixConciliationItem.created_at)
    )
    items = items_result.scalars().all()

    return {
        "run": ConciliacaoRunResponse.model_validate(run).model_dump(mode="json"),
        "items": [ConciliacaoItemResponse.model_validate(i).model_dump(mode="json") for i in items],
    }


@router.post("/conciliacao-pix/run")
async def force_conciliacao_pix(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Força nova execução de conciliação PIX para a data informada."""
    _require_financeiro(current_user)
    from app.services.financeiro.pix_conciliation_service import PixConciliationService

    raw_date = body.get("date")
    if raw_date:
        target = date.fromisoformat(str(raw_date))
    else:
        target = date.today()

    svc = PixConciliationService()
    data = await svc.run_daily_conciliation(db, target)
    run = data["run"]
    return ConciliacaoRunResponse.model_validate(run).model_dump(mode="json")


@router.post("/conciliacao-pix/{run_id}/aprovar")
async def aprovar_conciliacao(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Aprova fechamento de uma conciliação PIX."""
    _require_financeiro(current_user)
    from app.services.financeiro.pix_conciliation_service import PixConciliationService

    svc = PixConciliationService()
    try:
        run = await svc.aprovar_run(db, run_id, current_user.id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ConciliacaoRunResponse.model_validate(run).model_dump(mode="json")


@router.post("/conciliacao-pix/items/{item_id}/resolver")
async def resolver_item_conciliacao(
    item_id: UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Marca um item divergente da conciliação como resolvido."""
    _require_financeiro(current_user)
    from app.services.financeiro.pix_conciliation_service import PixConciliationService

    observacao = body.get("observacao", "")
    svc = PixConciliationService()
    try:
        item = await svc.resolver_item(db, item_id, current_user.id, observacao)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ConciliacaoItemResponse.model_validate(item).model_dump(mode="json")


@router.get("/conciliacao-pix/historico")
async def historico_conciliacao_pix(
    start: date = Query(...),
    end: date = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista runs de conciliação por período."""
    _require_financeiro(current_user)
    from app.models.financeiro.pix_conciliation import PixConciliationRun

    result = await db.execute(
        select(PixConciliationRun)
        .where(
            and_(
                PixConciliationRun.conciliation_date >= start,
                PixConciliationRun.conciliation_date <= end,
            )
        )
        .order_by(PixConciliationRun.conciliation_date.desc())
    )
    runs = result.scalars().all()
    return [ConciliacaoRunResponse.model_validate(r).model_dump(mode="json") for r in runs]


# ============================================================
# VASILHAMES / ESTOQUE
# ============================================================


class VasilhameEstoqueResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    tipo: str
    qtd_cheios: int
    qtd_vazios: int
    qtd_em_campo: int
    custo_unitario: Decimal
    valor_estoque: float


class VasilhameEntradaRequest(PydanticBaseModel):
    tipo: str
    quantidade: int
    custo_unitario: Decimal


class VasilhameRetornoRequest(PydanticBaseModel):
    tipo: str
    quantidade: int
    pedido_id: Optional[UUID] = None
    entregador_id: Optional[UUID] = None


class VasilhamePerdaRequest(PydanticBaseModel):
    tipo: str
    quantidade: int
    observacao: str


class VasilhameAjusteRequest(PydanticBaseModel):
    tipo: str
    qtd_cheios_real: int
    qtd_vazios_real: int
    observacao: str


@router.get("/vasilhames/posicao")
async def vasilhames_posicao(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna posição atual de estoque de todos os vasilhames."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        return await svc.get_posicao_atual(db)
    except ProgrammingError:
        return []


@router.get("/vasilhames/movimentacoes")
async def vasilhames_movimentacoes(
    tipo: Optional[str] = Query(None),
    start: Optional[date] = Query(None),
    end: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista paginada de movimentações de vasilhames."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        result = await svc.get_movimentacoes(db, tipo=tipo, start=start, end=end, page=page, limit=limit)
    except ProgrammingError:
        return {"total": 0, "page": page, "limit": limit, "items": []}
    items = result["items"]
    return {
        "total": result["total"],
        "page": result["page"],
        "limit": result["limit"],
        "items": [
            {
                "id": str(m.id),
                "tipo_vasilhame": m.tipo_vasilhame,
                "tipo_movimento": m.tipo_movimento,
                "quantidade": m.quantidade,
                "custo_unitario": float(m.custo_unitario) if m.custo_unitario else None,
                "custo_total": float(m.custo_total) if m.custo_total else None,
                "pedido_id": str(m.pedido_id) if m.pedido_id else None,
                "observacao": m.observacao,
                "saldo_cheios_antes": m.saldo_cheios_antes,
                "saldo_vazios_antes": m.saldo_vazios_antes,
                "saldo_em_campo_antes": m.saldo_em_campo_antes,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in items
        ],
    }


@router.post("/vasilhames/entrada", status_code=201)
async def vasilhames_entrada(
    body: VasilhameEntradaRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Registra entrada por compra de vasilhames."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        mov = await svc.registrar_entrada_compra(
            db,
            tipo=body.tipo,
            quantidade=body.quantidade,
            custo_unitario=body.custo_unitario,
            user_id=current_user.id,
        )
        await db.commit()
        return {"id": str(mov.id), "tipo_vasilhame": mov.tipo_vasilhame, "tipo_movimento": mov.tipo_movimento, "quantidade": mov.quantidade}
    except ProgrammingError:
        raise HTTPException(status_code=503, detail="Tabela de vasilhames não encontrada. Execute as migrações do banco de dados.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/vasilhames/retorno", status_code=201)
async def vasilhames_retorno(
    body: VasilhameRetornoRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Registra retorno de vasilhame vazio."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        mov = await svc.registrar_retorno_vazio(
            db,
            tipo=body.tipo,
            quantidade=body.quantidade,
            pedido_id=body.pedido_id,
            entregador_id=body.entregador_id,
            user_id=current_user.id,
        )
        await db.commit()
        return {"id": str(mov.id), "tipo_vasilhame": mov.tipo_vasilhame, "tipo_movimento": mov.tipo_movimento, "quantidade": mov.quantidade}
    except ProgrammingError:
        raise HTTPException(status_code=503, detail="Tabela de vasilhames não encontrada. Execute as migrações do banco de dados.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/vasilhames/perda", status_code=201)
async def vasilhames_perda(
    body: VasilhamePerdaRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Registra perda de vasilhame."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        mov = await svc.registrar_perda(
            db,
            tipo=body.tipo,
            quantidade=body.quantidade,
            observacao=body.observacao,
            user_id=current_user.id,
        )
        await db.commit()
        return {"id": str(mov.id), "tipo_vasilhame": mov.tipo_vasilhame, "tipo_movimento": mov.tipo_movimento, "quantidade": mov.quantidade}
    except ProgrammingError:
        raise HTTPException(status_code=503, detail="Tabela de vasilhames não encontrada. Execute as migrações do banco de dados.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/vasilhames/ajuste", status_code=201)
async def vasilhames_ajuste(
    body: VasilhameAjusteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ajuste de inventário de vasilhame."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        mov = await svc.ajuste_inventario(
            db,
            tipo=body.tipo,
            qtd_cheios_real=body.qtd_cheios_real,
            qtd_vazios_real=body.qtd_vazios_real,
            observacao=body.observacao,
            user_id=current_user.id,
        )
        await db.commit()
        return {"id": str(mov.id), "tipo_vasilhame": mov.tipo_vasilhame, "tipo_movimento": mov.tipo_movimento}
    except ProgrammingError:
        raise HTTPException(status_code=503, detail="Tabela de vasilhames não encontrada. Execute as migrações do banco de dados.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/vasilhames/alertas")
async def vasilhames_alertas(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna alertas de estoque baixo de vasilhames."""
    _require_financeiro(current_user)
    from app.services.financeiro.vasilhame_service import VasilhameService
    svc = VasilhameService()
    try:
        return await svc.get_alertas(db)
    except ProgrammingError:
        return []


# ============================================================
# FIADO — detalhamento e alertas
# ============================================================


class FiadoEntryResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    customer_id: UUID
    order_id: UUID
    valor_original: Decimal
    valor_pago: Decimal
    vencimento: date
    status: str
    cobrado_count: int
    pago_at: Optional[datetime]
    observacao: Optional[str]
    valor_pendente: Optional[Decimal] = None

    @classmethod
    def from_entry(cls, entry):
        return cls(
            id=entry.id,
            customer_id=entry.customer_id,
            order_id=entry.order_id,
            valor_original=entry.valor_original,
            valor_pago=entry.valor_pago,
            vencimento=entry.vencimento,
            status=entry.status,
            cobrado_count=entry.cobrado_count,
            pago_at=entry.pago_at,
            observacao=entry.observacao,
            valor_pendente=entry.valor_pendente,
        )


class FiadoPagamentoRequest(PydanticBaseModel):
    valor: Decimal = PydanticField(..., gt=0)
    metodo: str = PydanticField(default="dinheiro")
    asaas_payment_id: Optional[str] = None


@router.get("/fiado/aging")
async def fiado_aging(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Aging report de fiado: 0-7d, 8-15d, 16-30d, 30d+."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import get_aging_report
    return await get_aging_report(db)


@router.get("/fiado/entries")
async def list_fiado_entries(
    customer_id: Optional[UUID] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista paginada de entries de fiado."""
    _require_financeiro(current_user)
    from app.models.financeiro.fiado_entry import FiadoEntry
    import math

    filters = []
    if customer_id:
        filters.append(FiadoEntry.customer_id == customer_id)
    if status_filter:
        filters.append(FiadoEntry.status == status_filter)

    count_result = await db.execute(
        select(func.count(FiadoEntry.id)).where(and_(*filters) if filters else True)
    )
    total = count_result.scalar() or 0

    offset = (page - 1) * limit
    result = await db.execute(
        select(FiadoEntry)
        .where(and_(*filters) if filters else True)
        .order_by(FiadoEntry.vencimento.asc())
        .offset(offset)
        .limit(limit)
    )
    entries = result.scalars().all()

    return {
        "items": [FiadoEntryResponse.from_entry(e) for e in entries],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if limit else 1,
    }


@router.get("/fiado/entries/{entry_id}")
async def get_fiado_entry(
    entry_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Detalhe de uma entry com pagamentos."""
    _require_financeiro(current_user)
    from app.models.financeiro.fiado_entry import FiadoEntry
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(FiadoEntry)
        .options(selectinload(FiadoEntry.pagamentos))
        .where(FiadoEntry.id == entry_id)
    )
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(404, "Entry de fiado não encontrada")

    data = FiadoEntryResponse.from_entry(entry)
    data_dict = data.model_dump()
    data_dict["pagamentos"] = [
        {
            "id": str(p.id),
            "valor": float(p.valor),
            "metodo": p.metodo,
            "pago_at": p.pago_at.isoformat(),
            "asaas_payment_id": p.asaas_payment_id,
        }
        for p in entry.pagamentos
    ]
    return data_dict


@router.post("/fiado/entries/{entry_id}/pagar")
async def pagar_fiado_entry(
    entry_id: UUID,
    body: FiadoPagamentoRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Registra pagamento parcial ou total de uma entry de fiado."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import registrar_pagamento_fiado

    try:
        pagamento = await registrar_pagamento_fiado(
            db,
            entry_id=entry_id,
            valor=body.valor,
            metodo=body.metodo,
            user_id=current_user.id,
            asaas_payment_id=body.asaas_payment_id,
        )
        await db.commit()
        return {"id": str(pagamento.id), "valor": float(pagamento.valor), "metodo": pagamento.metodo}
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/fiado/entries/{entry_id}/cobrar-whatsapp")
async def cobrar_fiado_whatsapp(
    entry_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Envia cobrança de fiado via WhatsApp para o cliente."""
    _require_financeiro(current_user)
    from app.services.financeiro.customer_financial_service import enviar_cobranca_whatsapp

    ok = await enviar_cobranca_whatsapp(db, entry_id)
    await db.commit()
    return {"success": ok}


@router.post("/fiado/cobranca-lote")
async def cobranca_fiado_lote(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Envia cobrança em lote para todas as entries vencidas (admin/owner)."""
    if current_user.role not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Apenas admin/owner pode disparar cobrança em lote")
    from app.services.financeiro.customer_financial_service import (
        get_entries_para_cobranca,
        enviar_cobranca_whatsapp,
    )

    entries = await get_entries_para_cobranca(db)
    resultados = []
    for entry in entries:
        ok = await enviar_cobranca_whatsapp(db, entry.id)
        resultados.append({"entry_id": str(entry.id), "enviado": ok})

    await db.commit()
    return {"total_processados": len(resultados), "resultados": resultados}


# ============================================================
# NF-e
# ============================================================


class NotaFiscalResponse(PydanticBaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    tipo: str
    numero: Optional[int]
    serie: str
    chave_acesso: Optional[str]
    status: str
    pedido_id: Optional[UUID]
    customer_id: UUID
    valor_total: Decimal
    pdf_danfe_url: Optional[str]
    enviada_whatsapp: bool
    motivo_rejeicao: Optional[str]
    emitida_at: Optional[datetime]
    created_at: datetime


class EmitirNFeRequest(PydanticBaseModel):
    pedido_id: UUID


class CancelarNFeRequest(PydanticBaseModel):
    motivo: str = PydanticField(..., min_length=15)


@router.get("/nfe/resumo")
async def nfe_resumo(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Contagens de NF-e por status."""
    _require_financeiro(current_user)
    from app.models.financeiro.nota_fiscal import NotaFiscal

    result = await db.execute(
        select(NotaFiscal.status, func.count(NotaFiscal.id).label("total"))
        .group_by(NotaFiscal.status)
    )
    rows = result.all()
    return {r.status: r.total for r in rows}


@router.get("/nfe")
async def list_nfe(
    status_filter: Optional[str] = Query(None, alias="status"),
    start: Optional[date] = None,
    end: Optional[date] = None,
    customer_id: Optional[UUID] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista de notas fiscais com filtros."""
    _require_financeiro(current_user)
    from app.models.financeiro.nota_fiscal import NotaFiscal
    import math

    filters = []
    if status_filter:
        filters.append(NotaFiscal.status == status_filter)
    if start:
        filters.append(NotaFiscal.created_at >= start)
    if end:
        filters.append(NotaFiscal.created_at <= end)
    if customer_id:
        filters.append(NotaFiscal.customer_id == customer_id)

    count_result = await db.execute(
        select(func.count(NotaFiscal.id)).where(and_(*filters) if filters else True)
    )
    total = count_result.scalar() or 0

    offset = (page - 1) * limit
    result = await db.execute(
        select(NotaFiscal)
        .where(and_(*filters) if filters else True)
        .order_by(NotaFiscal.created_at.desc())
        .offset(offset)
        .limit(limit)
    )
    items = result.scalars().all()

    return {
        "items": [NotaFiscalResponse.model_validate(nf) for nf in items],
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if limit else 1,
    }


@router.get("/nfe/{nota_id}", response_model=NotaFiscalResponse)
async def get_nfe(
    nota_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Detalhe de uma nota fiscal."""
    _require_financeiro(current_user)
    from app.models.financeiro.nota_fiscal import NotaFiscal

    result = await db.execute(select(NotaFiscal).where(NotaFiscal.id == nota_id))
    nf = result.scalar_one_or_none()
    if not nf:
        raise HTTPException(404, "Nota fiscal não encontrada")
    return nf


@router.post("/nfe/emitir", response_model=NotaFiscalResponse, status_code=201)
async def emitir_nfe(
    body: EmitirNFeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Emissão manual de NF-e para um pedido."""
    _require_financeiro(current_user)
    from app.services.financeiro.nfe_service import NFeService

    try:
        service = NFeService()
        nf = await service.emitir_nfe_pedido(db, body.pedido_id, user_id=current_user.id)
        await db.commit()
        await db.refresh(nf)
        return nf
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/nfe/{nota_id}/cancelar", response_model=NotaFiscalResponse)
async def cancelar_nfe(
    nota_id: UUID,
    body: CancelarNFeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cancela uma NF-e autorizada."""
    if current_user.role not in ["admin", "owner"]:
        raise HTTPException(403, "Apenas admin/owner pode cancelar NF-e")
    from app.services.financeiro.nfe_service import NFeService

    try:
        service = NFeService()
        nf = await service.cancelar_nfe(db, nota_id, body.motivo, current_user.id)
        await db.commit()
        await db.refresh(nf)
        return nf
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/nfe/{nota_id}/reenviar-whatsapp")
async def reenviar_danfe_whatsapp(
    nota_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Reenvia DANFE via WhatsApp."""
    _require_financeiro(current_user)
    from app.services.financeiro.nfe_service import NFeService

    service = NFeService()
    ok = await service.reenviar_danfe(db, nota_id)
    await db.commit()
    return {"success": ok}


# ==================== DESPESAS ====================

from app.models.expense import Expense, ExpenseTipo, ExpenseTipoGrupo, Periodicidade
from app.schemas.expense import ExpenseCreate, ExpenseResponse, ExpensePagarBody, ExpensePorGrupoItem, ProximoVencimento
from app.services.expense_service import (
    create_expense,
    list_expenses,
    get_proximos_vencimentos,
    get_despesas_por_grupo,
    pagar_expense,
    pagar_grupo_parcela,
    delete_expense,
    get_alertas_vencimento,
)


@router.post("/despesas", response_model=List[ExpenseResponse], status_code=201)
async def criar_despesa(
    body: ExpenseCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cria despesa (única, parcelada ou recorrente)."""
    _require_financeiro(current_user)
    despesas = await create_expense(db, body, user_id=current_user.id)
    await db.commit()
    for d in despesas:
        await db.refresh(d)
    return despesas


@router.get("/despesas", response_model=List[ExpenseResponse])
async def listar_despesas(
    tipo: Optional[str] = Query(None),
    grupo: Optional[str] = Query(None),
    mes: Optional[int] = Query(None),
    ano: Optional[int] = Query(None),
    veiculo: Optional[str] = Query(None),
    funcionario_id: Optional[int] = Query(None),
    pago: Optional[bool] = Query(None),
    periodicidade: Optional[str] = Query(None),
    limit: int = Query(200, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista despesas com filtros."""
    _require_financeiro(current_user)
    return await list_expenses(
        db,
        tipo=tipo,
        grupo=grupo,
        mes=mes,
        ano=ano,
        veiculo=veiculo,
        funcionario_id=funcionario_id,
        pago=pago,
        periodicidade=periodicidade,
        limit=limit,
        offset=offset,
    )


@router.get("/despesas/proximos-vencimentos", response_model=List[ProximoVencimento])
async def proximos_vencimentos(
    dias: int = Query(30, ge=1, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Despesas não pagas vencendo nos próximos N dias."""
    _require_financeiro(current_user)
    from datetime import date as _date
    despesas = await get_proximos_vencimentos(db, dias=dias)
    hoje = _date.today()
    return [
        ProximoVencimento(
            id=d.id,
            tipo=d.tipo,
            grupo=d.grupo,
            descricao=d.descricao,
            valor_parcela=d.valor_parcela,
            data_vencimento=d.data_vencimento,
            dias_para_vencer=(d.data_vencimento - hoje).days,
            pago=d.pago,
            veiculo_placa=d.veiculo_placa,
        )
        for d in despesas
    ]


@router.get("/despesas/por-grupo", response_model=List[ExpensePorGrupoItem])
async def despesas_por_grupo(
    mes: Optional[int] = Query(None),
    ano: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Totais agrupados por grupo (para DRE)."""
    _require_financeiro(current_user)
    grupos = await get_despesas_por_grupo(db, mes=mes, ano=ano)
    return [ExpensePorGrupoItem(**g) for g in grupos]


@router.get("/despesas/alertas")
async def alertas_despesas(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Despesas urgentes: vencendo em 7 dias e já vencidas."""
    _require_financeiro(current_user)
    alertas = await get_alertas_vencimento(db)
    return {
        "vencendo_7_dias": [ExpenseResponse.model_validate(d) for d in alertas["vencendo_7_dias"]],
        "vencidas": [ExpenseResponse.model_validate(d) for d in alertas["vencidas"]],
        "total_urgente": len(alertas["vencendo_7_dias"]) + len(alertas["vencidas"]),
    }


@router.get("/despesas/por-veiculo/{placa}", response_model=List[ExpenseResponse])
async def despesas_por_veiculo(
    placa: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Todas as despesas de uma placa."""
    _require_financeiro(current_user)
    return await list_expenses(db, veiculo=placa)


@router.get("/despesas/por-funcionario/{funcionario_id}", response_model=List[ExpenseResponse])
async def despesas_por_funcionario(
    funcionario_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Todas as despesas de um funcionário."""
    _require_financeiro(current_user)
    return await list_expenses(db, funcionario_id=funcionario_id)


@router.put("/despesas/{expense_id}/pagar", response_model=ExpenseResponse)
async def pagar_despesa(
    expense_id: UUID,
    body: ExpensePagarBody = ExpensePagarBody(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Marca parcela como paga."""
    _require_financeiro(current_user)
    expense = await pagar_expense(db, expense_id, data_pagamento=body.data_pagamento)
    if not expense:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    await db.commit()
    await db.refresh(expense)
    return expense


@router.put("/despesas/grupo/{grupo_parcela_id}/pagar")
async def pagar_grupo_despesas(
    grupo_parcela_id: UUID,
    body: ExpensePagarBody = ExpensePagarBody(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Paga todas as parcelas de um grupo de uma vez."""
    _require_financeiro(current_user)
    count = await pagar_grupo_parcela(db, grupo_parcela_id, data_pagamento=body.data_pagamento)
    await db.commit()
    return {"pagas": count}


@router.delete("/despesas/{expense_id}", status_code=204)
async def deletar_despesa(
    expense_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cancela despesa (soft delete)."""
    _require_financeiro(current_user)
    ok = await delete_expense(db, expense_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    await db.commit()


# ─────────────────────────────────────────────────────────
# Aba Asaas — Consulta PIX
# ─────────────────────────────────────────────────────────

@router.get("/pix")
async def consultar_pix(
    date_from: Optional[date] = Query(None, description="Início do período (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, description="Fim do período (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Status: PENDING, RECEIVED, CONFIRMED, OVERDUE, REFUNDED"),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user),
):
    """
    Consulta cobranças PIX diretamente no Asaas.
    Aba exclusiva para Gracielle acompanhar pagamentos sem depender de grupos de WhatsApp.
    """
    _require_financeiro(current_user)
    from app.services.asaas_service import list_pix_payments

    result = await list_pix_payments(
        date_from=date_from,
        date_to=date_to,
        status=status,
        limit=limit,
        offset=offset,
    )

    # Normalise and enrich each payment for display
    payments = []
    for p in result.get("data", []):
        payments.append({
            "id": p.get("id"),
            "status": p.get("status"),
            "value": p.get("value"),
            "netValue": p.get("netValue"),
            "description": p.get("description"),
            "dueDate": p.get("dueDate"),
            "paymentDate": p.get("paymentDate"),
            "confirmedDate": p.get("confirmedDate"),
            "externalReference": p.get("externalReference"),
            "billingType": p.get("billingType"),
            "invoiceUrl": p.get("invoiceUrl"),
        })

    return {
        "data": payments,
        "totalCount": result.get("totalCount", len(payments)),
        "hasMore": result.get("hasMore", False),
        "offset": offset,
        "limit": limit,
    }


@router.get("/pix/exportar-excel")
async def exportar_pix_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    """
    Exporta cobranças PIX para Excel (.xlsx).
    Baixa todos os registros (até 500) no período selecionado.
    """
    _require_financeiro(current_user)
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.services.asaas_service import list_pix_payments

    result = await list_pix_payments(
        date_from=date_from,
        date_to=date_to,
        status=status,
        limit=500,
        offset=0,
    )
    payments = result.get("data", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PIX Asaas"

    # Header row
    headers = [
        "ID Asaas", "Status", "Valor (R$)", "Valor Líquido (R$)",
        "Descrição", "Vencimento", "Data Pagamento", "Referência Externa",
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    STATUS_PT = {
        "PENDING": "Pendente",
        "RECEIVED": "Recebido",
        "CONFIRMED": "Confirmado",
        "OVERDUE": "Vencido",
        "REFUNDED": "Estornado",
        "CANCELLED": "Cancelado",
    }

    for row_idx, p in enumerate(payments, start=2):
        ws.cell(row=row_idx, column=1, value=p.get("id", ""))
        ws.cell(row=row_idx, column=2, value=STATUS_PT.get(p.get("status", ""), p.get("status", "")))
        ws.cell(row=row_idx, column=3, value=p.get("value"))
        ws.cell(row=row_idx, column=4, value=p.get("netValue"))
        ws.cell(row=row_idx, column=5, value=p.get("description", ""))
        ws.cell(row=row_idx, column=6, value=p.get("dueDate", ""))
        ws.cell(row=row_idx, column=7, value=p.get("paymentDate") or p.get("confirmedDate") or "")
        ws.cell(row=row_idx, column=8, value=p.get("externalReference", ""))

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    periodo = ""
    if date_from:
        periodo += f"_{date_from}"
    if date_to:
        periodo += f"_{date_to}"
    filename = f"pix_asaas{periodo}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
